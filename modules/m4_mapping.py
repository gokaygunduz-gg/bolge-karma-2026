"""
m4_mapping.py — Kulüp → Şehir → Bölge Eşleştirme

Veri kaynağı: Kulüp Şehir Mapping.xlsx (Kulüp-Bölge-Şehir sayfası)
  Sütun A (Kulüp Alternatif) ve Sütun C (Kulüp Tekil) aranır.
  Sütun E → Şehir, Sütun G → Bölge (1-6)

⚠️  Bu modül ASLA Excel'e yazmaz. Sadece okur.
    Eşleşmeyen kulüpler log'a düşer, kullanıcıya bildirilir.
    Mapping güncellemesi Manuel olarak Excel'de yapılır.

Kullanım:
  from modules.m4_mapping import lookup_club, get_missing_clubs

  result = lookup_club("ANKARA BEL. SK")
  # → {"club_canonical": "Ankara ...", "city": "Ankara", "region": 4}
  # veya None (bulunamadı)

  missing = get_missing_clubs()
  # → ["YENİ KULÜP A.Ş.", "BİLİNMEYEN SK", ...]
"""

import openpyxl
import logging
from typing import TypedDict

from config import (
    MAPPING_EXCEL_PATH,
    MAPPING_SHEET_NAME,
    COL_CLUB_ALT,
    COL_CLUB_CANONICAL,
    COL_CITY,
    COL_REGION,
    MAPPING_DATA_START_ROW,
)
from modules.m1_normalize import normalize_for_lookup, restore_turkish_club

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# DB lookup yardımcısı (döngüsel import'u önlemek için lazy import)
# ─────────────────────────────────────────────────────────────────────────────

def _try_db_lookup(name_normalized: str) -> "ClubInfo | None":
    """
    DB'de arar. DB hazır değilse veya hata olursa None döner.
    Döngüsel import'u önlemek için içeride import yapılır.
    """
    try:
        from database.db import lookup_club_db
        row = lookup_club_db(name_normalized)
        if row:
            return ClubInfo(
                club_canonical = restore_turkish_club(row["name_canonical"] or name_normalized),
                city           = row["city"],
                region         = row["region"],
            )
    except Exception:
        pass  # DB erişimi başarısız → Excel cache'e düş
    return None


def _try_db_lookup_nospace(name_normalized_nospace: str) -> "ClubInfo | None":
    """OCR birleşik kulüp adları için boşluk kaldırılmış anahtar ile DB'de arar."""
    try:
        from database.db import lookup_club_db_nospace
        row = lookup_club_db_nospace(name_normalized_nospace)
        if row:
            return ClubInfo(
                club_canonical = restore_turkish_club(row["name_canonical"] or name_normalized_nospace),
                city           = row["city"],
                region         = row["region"],
            )
    except Exception:
        pass
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Tip tanımı
# ─────────────────────────────────────────────────────────────────────────────

class ClubInfo(TypedDict):
    club_canonical: str   # Excel'deki standart kulüp adı (Sütun C, yoksa A)
    city:           str   # Şehir (Sütun E)
    region:         int   # Bölge (Sütun G, 1-6)


# ─────────────────────────────────────────────────────────────────────────────
# 1. Mapping yükleme (ilk çağrıda bir kez yüklenir, cache'de tutulur)
# ─────────────────────────────────────────────────────────────────────────────

# _mapping_cache: normalize_for_lookup(kulüp_adı) → ClubInfo
_mapping_cache: dict[str, ClubInfo] | None = None

# Bulunamayan kulüpler kümesi (tekrar aramayı önler, kullanıcıya toplu bildirim)
# _mapping_nospace_cache: bosluk_kaldirilmis → ClubInfo (OCR fallback)
_mapping_nospace_cache: dict[str, ClubInfo] | None = None

_missing_clubs: set[str] = set()


def _load_mapping() -> dict[str, ClubInfo]:
    """
    Excel'i okur ve normalize edilmiş anahtar → ClubInfo sözlüğü oluşturur.

    Her satır için iki anahtar ekler:
      - normalize_for_lookup(col_A)  (alternatif isim)
      - normalize_for_lookup(col_C)  (kanonik isim, varsa)
    İki anahtar da aynı ClubInfo'ya işaret eder.

    Satır atlanır:
      - Hem A hem C sütunu boşsa
      - Şehir veya Bölge boşsa
    """
    mapping: dict[str, ClubInfo] = {}

    try:
        wb = openpyxl.load_workbook(MAPPING_EXCEL_PATH, read_only=True, data_only=True)
    except FileNotFoundError:
        logger.error("Mapping Excel bulunamadı: %s", MAPPING_EXCEL_PATH)
        return mapping
    except Exception as e:
        logger.error("Mapping Excel açılamadı: %s", e)
        return mapping

    if MAPPING_SHEET_NAME not in wb.sheetnames:
        logger.error("Sheet bulunamadı: '%s'. Mevcut sheetler: %s",
                     MAPPING_SHEET_NAME, wb.sheetnames)
        wb.close()
        return mapping

    ws = wb[MAPPING_SHEET_NAME]
    rows_loaded = 0
    rows_skipped = 0

    for row in ws.iter_rows(min_row=MAPPING_DATA_START_ROW, values_only=True):
        col_a   = row[COL_CLUB_ALT]        if len(row) > COL_CLUB_ALT        else None
        col_c   = row[COL_CLUB_CANONICAL]  if len(row) > COL_CLUB_CANONICAL  else None
        col_e   = row[COL_CITY]            if len(row) > COL_CITY            else None
        col_g   = row[COL_REGION]          if len(row) > COL_REGION          else None

        # Şehir veya Bölge yoksa bu satırı atla (veri eksik)
        if not col_e or col_g is None:
            rows_skipped += 1
            continue

        # En az bir kulüp ismi olmalı
        if not col_a and not col_c:
            rows_skipped += 1
            continue

        # Kanonik isim: Önce C sütunu, yoksa A sütunu — Türkçe karakter düzeltmesi uygula
        canonical = restore_turkish_club(str(col_c).strip() if col_c else str(col_a).strip())
        city      = str(col_e).strip()

        try:
            region = int(col_g)
        except (ValueError, TypeError):
            rows_skipped += 1
            continue

        info: ClubInfo = {
            "club_canonical": canonical,
            "city":           city,
            "region":         region,
        }

        # Alternatif isim (A sütunu) → mapping'e ekle
        if col_a:
            key_a = normalize_for_lookup(str(col_a))
            if key_a and key_a not in mapping:
                mapping[key_a] = info

        # Kanonik isim (C sütunu) → mapping'e ekle
        if col_c:
            key_c = normalize_for_lookup(str(col_c))
            if key_c and key_c not in mapping:
                mapping[key_c] = info

        rows_loaded += 1

    wb.close()
    logger.info("Mapping yüklendi: %d kulüp, %d satır atlandı.", rows_loaded, rows_skipped)
    return mapping


def _get_mapping() -> dict[str, ClubInfo]:
    """Singleton pattern: Mapping'i ilk çağrıda yükle, sonra cache'den sun."""
    global _mapping_cache
    if _mapping_cache is None:
        _mapping_cache = _load_mapping()
    return _mapping_cache


def _get_mapping_nospace() -> dict[str, ClubInfo]:
    """Boşluk kaldırılmış normalize anahtarlar → ClubInfo (OCR fallback cache)."""
    global _mapping_nospace_cache
    if _mapping_nospace_cache is None:
        _mapping_nospace_cache = {
            k.replace(" ", ""): v
            for k, v in _get_mapping().items()
        }
    return _mapping_nospace_cache


def reload_mapping() -> None:
    """
    Mapping'i yeniden yükle (Excel güncellendikten sonra kullanılır).
    Normal kullanımda çağrılması gerekmez.
    """
    global _mapping_cache, _mapping_nospace_cache, _missing_clubs
    _mapping_cache = None
    _mapping_nospace_cache = None
    _missing_clubs.clear()
    logger.info("Mapping cache temizlendi, yeniden yuklenecek.")


def apply_overrides_to_mapping(override_path: str | None = None) -> int:
    """
    manual_overrides.json içindeki club_aliases'ı bellek cache'ine ekler.
    reload_mapping()'den SONRA çağrılmalıdır.

    Döndürür: eklenen alias sayısı
    """
    import json, os
    from pathlib import Path

    if override_path is None:
        override_path = str(Path(__file__).parent.parent / "manual_overrides.json")

    if not os.path.exists(override_path):
        return 0

    try:
        with open(override_path, encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        logger.warning("manual_overrides.json okunamadı: %s", e)
        return 0

    # Cache'in yüklenmiş olduğundan emin ol
    mapping = _get_mapping()

    added = 0
    for alias in data.get("club_aliases", []):
        raw = alias.get("raw", "").strip()
        if not raw or raw.startswith("_") or alias.get("_ornek"):
            continue
        canonical = alias.get("canonical") or raw
        city      = alias.get("city", "")
        region    = alias.get("region", 0)
        if not city or not region:
            continue
        info: ClubInfo = {
            "club_canonical": canonical,
            "city":           city,
            "region":         int(region),
        }
        key = normalize_for_lookup(raw)
        if key and key not in mapping:
            mapping[key] = info
            # nospace cache de güncelle
            ns = _get_mapping_nospace()
            ns[key.replace(" ", "")] = info
            added += 1
            logger.debug("Override alias eklendi: %s → %s (%s, B%d)", raw, canonical, city, region)

    if added:
        logger.info("manual_overrides.json: %d yeni club alias eklendi.", added)
    return added


# ─────────────────────────────────────────────────────────────────────────────
# 2. Arama fonksiyonları
# ─────────────────────────────────────────────────────────────────────────────

def lookup_club(club_name: str | None) -> ClubInfo | None:
    """
    Kulüp adını arar, Şehir ve Bölge bilgisini döndürür.

    Arama sırası:
      1. SQLite DB (en güncel — Excel sync sonrası)
      2. Excel bellek cache (DB erişilemezse fallback)
      3. Eşleşme yoksa → None döner, kulüp _missing_clubs setine eklenir

    Parametreler:
      club_name: Yarış sonucundan gelen kulüp adı (her türlü format kabul)

    Döndürür:
      ClubInfo (club_canonical, city, region) veya None

    Örnek:
      lookup_club("ANKARA BEL. SK")
      → {"club_canonical": "...", "city": "Ankara", "region": 4}

      lookup_club("BİLİNMEYEN KULÜP")
      → None  (ve "BİLİNMEYEN KULÜP" missing listesine eklenir)
    """
    if not club_name or not club_name.strip():
        return None

    club_name = club_name.strip()
    normalized = normalize_for_lookup(club_name)

    if not normalized:
        return None

    # 1. DB'de ara (tam normalize eslesme)
    result = _try_db_lookup(normalized)

    # 2. DB'de yoksa Excel cache'e bak (tam normalize eslesme)
    if result is None:
        mapping = _get_mapping()
        result  = mapping.get(normalized)

    # 3. OCR fallback: bosluk kaldirilmis normalize anahtarla yeniden dene
    # (hem bosluklu hem bosluksuz OCR varyantlari icin — kos ici gecersiz)
    if result is None:
        normalized_nospace = normalized.replace(" ", "")
        result = _get_mapping_nospace().get(normalized_nospace)
        if result is None:
            result = _try_db_lookup_nospace(normalized_nospace)

    if result is None:
        _missing_clubs.add(club_name)
        logger.debug("Kulup bulunamadi: '%s' (normalize: '%s')", club_name, normalized)

    return result


def lookup_clubs_batch(club_names: list[str]) -> dict[str, ClubInfo | None]:
    """
    Birden fazla kulüp adını toplu olarak arar.

    Döndürür:
      {club_name: ClubInfo | None, ...}

    Örnek:
      lookup_clubs_batch(["Fenerbahçe SK", "Bilinmeyen Kulüp"])
      → {"Fenerbahçe SK": {...}, "Bilinmeyen Kulüp": None}
    """
    return {name: lookup_club(name) for name in club_names}


# ─────────────────────────────────────────────────────────────────────────────
# 3. Eksik kulüp yönetimi
# ─────────────────────────────────────────────────────────────────────────────

def get_missing_clubs() -> list[str]:
    """
    Bu oturumda bulunamayan kulüplerin listesini döndürür.
    Kullanıcıya gösterilmeli; bu kulüplerin Excel'e eklenmesi gerekir.

    Döndürür:
      Alfabetik sıralı, benzersiz kulüp adları listesi
    """
    return sorted(_missing_clubs)


def has_missing_clubs() -> bool:
    """Eşleşmeyen kulüp var mı?"""
    return len(_missing_clubs) > 0


def clear_missing_clubs() -> None:
    """Missing log'unu temizle (genellikle yeni bir yarış analizi başlangıcında)."""
    _missing_clubs.clear()


def report_missing_clubs() -> str:
    """
    Bulunamayan kulüplerin okunabilir raporunu döndürür.
    Dashboard veya terminal çıktısı için kullanılır.
    """
    missing = get_missing_clubs()
    if not missing:
        return "✓ Tüm kulüpler eşleşti."

    lines = [
        f"⚠️  {len(missing)} kulüp mapping'de bulunamadı.",
        "Bu kulüplerin Excel'e eklenmesi gerekiyor:",
        "",
    ]
    for i, club in enumerate(missing, 1):
        lines.append(f"  {i:3}. {club}")
    lines.append("")
    lines.append("Excel: Kulüp Şehir Mapping Exceli/Kulüp Şehir Mapping.xlsx")
    lines.append("Ekledikten sonra reload_mapping() çağırın.")
    return "\n".join(lines)


# ─────────────────────────────────────────────────────────────────────────────
# 4. Mapping istatistikleri (debug/bilgi amaçlı)
# ─────────────────────────────────────────────────────────────────────────────

def mapping_stats() -> dict:
    """
    Yüklenen mapping hakkında özet bilgi döndürür.
    Kullanım: Sistem başlangıcında veya debug için.
    """
    mapping = _get_mapping()
    if not mapping:
        return {"loaded": False, "total_keys": 0}

    from collections import Counter
    region_counts = Counter(info["region"] for info in mapping.values())

    return {
        "loaded":       True,
        "total_keys":   len(mapping),
        "by_region":    dict(sorted(region_counts.items())),
    }
