"""
compare_antalya_sources.py
--------------------------
CS-371 Antalya 3 kaynağını ayrı ayrı çeker, karşılaştırır, Excel'e atar.

KAYNAK 0: Lenex XML   — https://canli.tyf.gov.tr/tyf/cs-371/
KAYNAK A: Bireysel PDF — ResultList_1..50.pdf (canli.tyf.gov.tr)
KAYNAK B: Toplu PDF   — dosya.tyf.gov.tr (OCR ~10 dk)

Her kaynaktan ayrı Excel üretilir; karşılaştırma tablosu yazılır.
Ayrıca sıralama/federasyon karma tablosu da üretilir.
"""

import sys, os, time
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import logging
logging.basicConfig(level=logging.WARNING, format="%(levelname)s | %(name)s | %(message)s")

from database.db import init_db
from modules.m2_scraper import (
    scrape_direct_pdf, _enrich_all, _merge_abbreviated_names,
    _dedup_best_time, _add_race_date, _print_summary,
)
from modules.m4_mapping import clear_missing_clubs, reload_mapping
from parsers.lenex_parser import download_lenex, parse_lenex, parse_lenex_date
from parsers.pdf_parser import parse_pdf_from_url
from export.excel_exporter import export_race_results

BASE_URL  = "https://canli.tyf.gov.tr/tyf/cs-371/"
PDF_BASE  = "https://canli.tyf.gov.tr/tyf/cs-371/canli/"
TOPLU_URL = "https://dosya.tyf.gov.tr/public/upload/0/2025-12/1766426510.pdf"
RACE_DATE = "2025.12.20"
TITLE     = "12-13 Yaş Türkiye Finali (Antalya)"

OUT_DIR   = os.path.join(os.path.dirname(__file__),
                         "Çıktılar", "Yarış Sonuçları Çıktı")
KARMA_DIR = os.path.join(os.path.dirname(__file__),
                         "Çıktılar", "Bölge Karmaları Sonuç Çıktı")
os.makedirs(OUT_DIR,   exist_ok=True)
os.makedirs(KARMA_DIR, exist_ok=True)


# ─────────────────────────────────────────────────────────────────────────────
def _stats(results):
    if not results:
        return 0, 0, 0
    total  = len(results)
    found  = sum(1 for r in results if r.get("region"))
    names  = len(set(r["name"] for r in results if r.get("name")))
    return total, names, round(100 * found / total) if total else 0


# ─────────────────────────────────────────────────────────────────────────────
# KAYNAK 0 — Lenex
# ─────────────────────────────────────────────────────────────────────────────
def kaynak_lenex():
    print("\n" + "═"*60)
    print("KAYNAK 0: LENEX XML")
    print("═"*60)
    t0 = time.time()
    lenex_content = download_lenex(BASE_URL)
    if not lenex_content:
        print("  ✗ Lenex bulunamadı (yarış bitmeden ya da dosya kaldırılmış)")
        return []
    race_date = parse_lenex_date(lenex_content) or RACE_DATE
    raw = parse_lenex(lenex_content)
    if not raw:
        print("  ✗ Lenex parse edildi ama sonuç yok")
        return []
    results = _enrich_all(raw)
    results = _merge_abbreviated_names(results)
    results = _dedup_best_time(results)
    _add_race_date(results, race_date)
    total, names, pct = _stats(results)
    print(f"  ✓ {total} satır, {names} sporcu, {pct}% kulüp eşleşmesi  ({time.time()-t0:.1f}s)")
    return results


# ─────────────────────────────────────────────────────────────────────────────
# KAYNAK A — Bireysel ResultList PDF'ler
# ─────────────────────────────────────────────────────────────────────────────
def kaynak_bireysel_pdf():
    print("\n" + "═"*60)
    print("KAYNAK A: BİREYSEL ResultList PDF'LERİ")
    print("═"*60)
    t0 = time.time()

    clear_missing_clubs()
    raw_results = []
    found = 0
    consecutive_empty = 0

    for n in range(1, 60):
        url = f"{PDF_BASE}ResultList_{n}.pdf"
        res = parse_pdf_from_url(url)
        if res:
            raw_results.extend(res)
            found += 1
            consecutive_empty = 0
            print(f"  ✓ ResultList_{n:02d}.pdf → {len(res)} satır")
        else:
            consecutive_empty += 1
            if consecutive_empty == 1:
                print(f"  ✗ ResultList_{n:02d}.pdf → boş/yok")
            elif consecutive_empty == 5:
                print(f"  (5 ardışık boş — daha fazla denenmeyecek)")
                break

    print(f"\n  Çekilen: {found} PDF")
    if not raw_results:
        return []

    results = _enrich_all(raw_results)
    results = _merge_abbreviated_names(results)
    results = _dedup_best_time(results)
    _add_race_date(results, RACE_DATE)
    total, names, pct = _stats(results)
    print(f"  ✓ {total} satır, {names} sporcu, {pct}% kulüp eşleşmesi  ({time.time()-t0:.1f}s)")
    return results


# ─────────────────────────────────────────────────────────────────────────────
# KAYNAK B — Toplu PDF (OCR)
# ─────────────────────────────────────────────────────────────────────────────
def kaynak_toplu_pdf():
    print("\n" + "═"*60)
    print("KAYNAK B: TOPLU PDF (OCR — bu adım ~5-10 dakika sürebilir)")
    print("═"*60)
    print(f"  URL: {TOPLU_URL}")
    t0 = time.time()
    results = scrape_direct_pdf(TOPLU_URL, verbose=True)
    if results:
        _add_race_date(results, RACE_DATE)
    total, names, pct = _stats(results)
    print(f"  ✓ {total} satır, {names} sporcu, {pct}% kulüp eşleşmesi  ({time.time()-t0:.1f}s)")
    return results


# ─────────────────────────────────────────────────────────────────────────────
# Excel export yardımcısı
# ─────────────────────────────────────────────────────────────────────────────
def export(results, suffix, source_label):
    if not results:
        print(f"  ✗ {source_label}: sonuç yok, Excel yazılmadı")
        return None
    fname   = f"2025.12.20_cs371_Antalya_{suffix}.xlsx"
    fpath   = os.path.join(OUT_DIR, fname)
    path    = export_race_results(results=results, filename=fname, title=f"{TITLE} — {source_label}")
    print(f"  Excel: {path}")
    return path


# ─────────────────────────────────────────────────────────────────────────────
# Karşılaştırma tablosu
# ─────────────────────────────────────────────────────────────────────────────
def compare(results_map: dict):
    print("\n" + "═"*60)
    print("KARŞILAŞTIRMA")
    print("═"*60)
    print(f"  {'Kaynak':<22} {'Satır':>6} {'Sporcu':>7} {'Kulüp%':>7}")
    print("  " + "─"*46)

    for label, results in results_map.items():
        if not results:
            print(f"  {label:<22} {'—':>6} {'—':>7} {'—':>7}")
            continue
        total, names, pct = _stats(results)
        print(f"  {label:<22} {total:>6} {names:>7} {pct:>6}%")

    # Sadece puanlı sporcularda (2011/2012/2013) fark
    print()
    target_bys = {2013, 2012, 2011}
    for label, results in results_map.items():
        if not results:
            continue
        karma_athletes = set()
        for r in results:
            by = r.get("birth_year")
            if by in target_bys:
                karma_athletes.add(r.get("name", ""))
        print(f"  {label:<22}: karma yaş grubu sporcu = {len(karma_athletes)}")


# ─────────────────────────────────────────────────────────────────────────────
# Federasyon karma sıralaması Excel'e yaz
# ─────────────────────────────────────────────────────────────────────────────
def export_karma_rankings():
    print("\n" + "═"*60)
    print("FED. KARMA SIRALAMASI -> Excel")
    print("═"*60)
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from federasyon.db_fed import load_athletes_for_ranking, load_athletes_for_ranking_by_leg, init_fed_db
        from federasyon.ranker import rank_all, format_seq_display, compute_club_rankings, compute_summary_matrix
        from federasyon.scoring_tables import SELECTION_QUOTAS
        from federasyon.scorer import best_scores_sequence

        init_fed_db()
        athletes    = load_athletes_for_ranking([2013, 2012, 2011])
        by_leg_data = load_athletes_for_ranking_by_leg([2013, 2012, 2011])
        if not athletes:
            print("  ✗ DB bos! Once process_antalya_karma.py calistirin.")
            return

        ranked = rank_all(athletes)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        AGE_LABEL    = {2013: "13 Yas (2013)", 2012: "14 Yas (2012)", 2011: "15 Yas (2011)"}
        GENDER_LABEL = {"F": "Kadin", "M": "Erkek"}
        SEL_LABEL    = {
            "TR": "TR Kadro", "BÖLGE": "Bölge", "BARAJ_YOK": "Baraj",
            "MULTINATIONS": "Multinations", "-": "—"
        }

        # Renkler
        CLR_MULTI = PatternFill("solid", fgColor="7B4F00")   # koyu altin - Multinations
        CLR_TR    = PatternFill("solid", fgColor="1B5E20")   # koyu yesil - TR
        CLR_BOLGE = PatternFill("solid", fgColor="0D47A1")   # koyu mavi - Bolge
        CLR_TIE   = PatternFill("solid", fgColor="4A235A")   # mor - esit
        CLR_HDR   = PatternFill("solid", fgColor="1A237E")
        CLR_SUBHDR= PatternFill("solid", fgColor="263238")
        FONT_W    = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
        FONT_NORM = Font(color="1A1A1A", name="Calibri", size=10)   # koyu renk — beyaz zemin üzerinde görünür
        ALIGN_C   = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ALIGN_L   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

        thin = Side(style="thin", color="37474F")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        EVENT_COLS = [
            ("Serbest",50), ("Serbest",100), ("Serbest",200), ("Serbest",400),
            ("Serbest",800), ("Serbest",1500),
            ("Sirtüstü",50), ("Sirtüstü",100), ("Sirtüstü",200),
            ("Kurbagalama",50), ("Kurbagalama",100), ("Kurbagalama",200),
            ("Kelebek",50), ("Kelebek",100), ("Kelebek",200),
            ("Karisik",200), ("Karisik",400),
        ]
        # Gerçek DB anahtarları (Turkish chars)
        EVENT_COLS_DB = [
            ("Serbest",50), ("Serbest",100), ("Serbest",200), ("Serbest",400),
            ("Serbest",800), ("Serbest",1500),
            ("Sırtüstü",50), ("Sırtüstü",100), ("Sırtüstü",200),
            ("Kurbağalama",50), ("Kurbağalama",100), ("Kurbağalama",200),
            ("Kelebek",50), ("Kelebek",100), ("Kelebek",200),
            ("Karışık",200), ("Karışık",400),
        ]

        def write_group_sheet(wb, sheet_name, group, quota, by_leg):
            ws = wb.create_sheet(title=sheet_name[:31])
            ws.freeze_panes = "A3"

            # Sütun sayısı: BASE + Antalya puan + Edirne puan + Birlikte puan
            # BASE: Sıra, Ad, Kulüp, Şehir, Blg, Seçim, Eşit, Top3(A), Top3(E), Top3(A+E), Dizi, Blg Sıra
            N_BASE = 12
            N_EV   = len(EVENT_COLS)
            N_TOTAL= N_BASE + N_EV * 3  # antalya + edirne + combined

            last_col = openpyxl.utils.get_column_letter(N_TOTAL)
            ws.merge_cells(f"A1:{last_col}1")
            hdr = ws["A1"]
            hdr.value = (f"2026 Federasyon Karması — {sheet_name} "
                         f"| TR: {quota.get('tr',0)} | B1: {quota.get('region_1',0)} | Diger: {quota.get('region_other',0)}"
                         f" | Baraj: {quota.get('min_points',7)} puan")
            hdr.fill = CLR_HDR; hdr.font = FONT_W; hdr.alignment = ALIGN_C

            BASE_HDRS = [
                "Sira", "Ad Soyad", "Kulüp", "Şehir", "Blg", "Seçim", "Esit?",
                "Top3\nAntalya", "Top3\nEdirne", "Top3\nA+E", "Dizi (A+E)", "Blg Sira"
            ]
            EV_HDRS_A = [f"{d}m {s[:3]}\n(Antalya)" for s, d in EVENT_COLS]
            EV_HDRS_E = [f"{d}m {s[:3]}\n(Edirne)"  for s, d in EVENT_COLS]
            EV_HDRS_C = [f"{d}m {s[:3]}\n(A+E)"     for s, d in EVENT_COLS]
            all_hdrs = BASE_HDRS + EV_HDRS_A + EV_HDRS_E + EV_HDRS_C

            for ci, h in enumerate(all_hdrs, 1):
                cell = ws.cell(row=2, column=ci, value=h)
                cell.fill = CLR_SUBHDR; cell.font = FONT_W
                cell.alignment = ALIGN_C; cell.border = border

            widths = [6, 26, 26, 12, 5, 14, 5, 7, 7, 7, 16, 6]
            widths += [5] * (N_EV * 3)
            for ci, w in enumerate(widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
            ws.row_dimensions[2].height = 32

            for ri, a in enumerate(group, 3):
                sel = a.get("selected", "-")
                tied = a.get("tied", False)

                if sel == "MULTINATIONS":   fill = CLR_MULTI
                elif tied and sel in ("-", "BÖLGE"): fill = CLR_TIE
                elif sel == "TR":           fill = CLR_TR
                elif sel == "BÖLGE":        fill = CLR_BOLGE
                else:                       fill = None

                key = (a["name"], a["birth_year"])
                leg = by_leg.get(key, {})
                ant_scores = leg.get("antalya", {})
                edr_scores = leg.get("edirne",  {})
                comb_scores= a.get("event_scores", {})

                top3_ant = sum(best_scores_sequence(ant_scores)[:3]) if ant_scores else 0
                top3_edr = sum(best_scores_sequence(edr_scores)[:3])  if edr_scores  else 0
                top3_comb= sum(a.get("seq",[])[:3])

                row_data = [
                    "★" if sel == "MULTINATIONS" else (a.get("tr_rank") or ""),
                    a["name"] + (" ★" if tied else ""),
                    a.get("club", ""),
                    a.get("city", ""),
                    a.get("region", ""),
                    SEL_LABEL.get(sel, sel),
                    "Evet" if tied else "",
                    top3_ant,
                    top3_edr,
                    top3_comb,
                    format_seq_display(a.get("seq", [])),
                    a.get("region_rank", "") or "",
                ]
                for ev in EVENT_COLS_DB:
                    row_data.append(ant_scores.get(ev, ""))
                for ev in EVENT_COLS_DB:
                    row_data.append(edr_scores.get(ev, ""))
                for ev in EVENT_COLS_DB:
                    row_data.append(comb_scores.get(ev, ""))

                for ci, val in enumerate(row_data, 1):
                    cell = ws.cell(row=ri, column=ci, value=val)
                    if fill:
                        cell.fill = fill
                        cell.font = FONT_W
                    else:
                        cell.font = FONT_NORM
                    cell.alignment = ALIGN_C if ci != 2 else ALIGN_L
                    cell.border = border

        from itertools import groupby as igrp
        ranked_sorted = sorted(ranked, key=lambda a: (a["birth_year"], a["gender"]))

        for (by, gender), grp_iter in igrp(ranked_sorted, key=lambda a: (a["birth_year"], a["gender"])):
            group = list(grp_iter)
            quota = SELECTION_QUOTAS.get(by, {})
            sheet_name = f"{AGE_LABEL[by]} {GENDER_LABEL[gender]}"
            write_group_sheet(wb, sheet_name, group, quota, by_leg_data)

        # ── Kulüp Sıralaması sayfası ──────────────────────────────────────────
        club_rankings = compute_club_rankings(ranked)
        ws_club = wb.create_sheet(title="Kulüp Sıralaması")
        ws_club.freeze_panes = "A3"
        ws_club.merge_cells("A1:I1")
        c = ws_club["A1"]
        c.value = "2026 Federasyon Karması — Kulüp Sıralaması (Multinations + TR Kadro + Bölge)"
        c.fill = CLR_HDR; c.font = FONT_W; c.alignment = ALIGN_C

        GROUP_ORDER = ["overall","2013_F","2013_M","2012_F","2012_M","2011_F","2011_M"]
        GROUP_LABEL = {
            "overall":"Genel","2013_F":"13K","2013_M":"13E",
            "2012_F":"14K","2012_M":"14E","2011_F":"15K","2011_M":"15E"
        }
        club_hdrs = ["#","Kulüp"] + [GROUP_LABEL[g] for g in GROUP_ORDER]
        for ci, h in enumerate(club_hdrs, 1):
            cell = ws_club.cell(row=2, column=ci, value=h)
            cell.fill = CLR_SUBHDR; cell.font = FONT_W
            cell.alignment = ALIGN_C; cell.border = border
        ws_club.column_dimensions["A"].width = 4
        ws_club.column_dimensions["B"].width = 34
        for ci in range(3, 3+len(GROUP_ORDER)):
            ws_club.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 8

        # overall sıralamasına göre kulüpleri yaz
        all_clubs = {c["club"]: c for c in club_rankings["overall"]}
        ri = 3
        for c_overall in club_rankings["overall"]:
            club = c_overall["club"]
            if c_overall["total"] == 0:
                continue
            row_data = [ri-2, club]
            for gk in GROUP_ORDER:
                if gk == "overall":
                    row_data.append(f"M{c_overall['multi']} T{c_overall['tr']} B{c_overall['bolge']}={c_overall['total']}")
                else:
                    grp_data = next((x for x in club_rankings["by_group"].get(gk,[]) if x["club"]==club), None)
                    if grp_data and grp_data["total"] > 0:
                        row_data.append(f"M{grp_data['multi']} T{grp_data['tr']} B{grp_data['bolge']}={grp_data['total']}")
                    else:
                        row_data.append("")
            for ci, val in enumerate(row_data, 1):
                cell = ws_club.cell(row=ri, column=ci, value=val)
                cell.font = FONT_NORM
                cell.alignment = ALIGN_C if ci != 2 else ALIGN_L
                cell.border = border
            ri += 1

        # ── Özet Matris sayfası ───────────────────────────────────────────────
        summary = compute_summary_matrix(ranked)
        ws_mat = wb.create_sheet(title="Özet Matris")
        ws_mat.merge_cells("A1:F1")
        c2 = ws_mat["A1"]
        c2.value = "2026 Federasyon Karması — Seçim Özeti"
        c2.fill = CLR_HDR; c2.font = FONT_W; c2.alignment = ALIGN_C
        mat_hdrs = ["Grup", "Multinations", "TR Kadro", "Bölge", "Toplam", "Not"]
        for ci, h in enumerate(mat_hdrs, 1):
            cell = ws_mat.cell(row=2, column=ci, value=h)
            cell.fill = CLR_SUBHDR; cell.font = FONT_W
            cell.alignment = ALIGN_C; cell.border = border
        ws_mat.column_dimensions["A"].width = 18
        ws_mat.column_dimensions["B"].width = 14
        ws_mat.column_dimensions["C"].width = 10
        ws_mat.column_dimensions["D"].width = 8
        ws_mat.column_dimensions["E"].width = 8
        ws_mat.column_dimensions["F"].width = 40
        AGE_L = {2013:"13 Yas",2012:"14 Yas",2011:"15 Yas"}
        GEN_L = {"F":"Kadin","M":"Erkek"}
        t_m=t_t=t_b=t_all=0
        for ri, g in enumerate(summary, 3):
            label = f"{AGE_L.get(g['birth_year'],g['birth_year'])} {GEN_L.get(g['gender'],g['gender'])}"
            t_m+=g['multi']; t_t+=g['tr']; t_b+=g['bolge']; t_all+=g['total']
            for ci, val in enumerate([label, g['multi'], g['tr'], g['bolge'], g['total'], ""], 1):
                cell = ws_mat.cell(row=ri, column=ci, value=val)
                cell.font = FONT_NORM; cell.alignment = ALIGN_C if ci>1 else ALIGN_L
                cell.border = border
        for ci, val in enumerate(["TOPLAM", t_m, t_t, t_b, t_all, ""], 1):
            cell = ws_mat.cell(row=len(summary)+3, column=ci, value=val)
            cell.font = FONT_W; cell.alignment = ALIGN_C if ci>1 else ALIGN_L
            cell.border = border

        legs_label = "Sadece Antalya"
        fname = f"2025.12.20_Antalya_FedKarma_Siralamasi.xlsx"
        fpath = os.path.join(KARMA_DIR, fname)
        wb.save(fpath)
        print(f"  ✓ Siralama Excel: {fpath}")
        print(f"  Sayfalar: {[s.title for s in wb.worksheets]}")

    except Exception as e:
        import traceback
        print(f"  ✗ Karma Excel hatası: {e}")
        traceback.print_exc()


# ─────────────────────────────────────────────────────────────────────────────
# ANA AKIŞ
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys
    skip_ocr = "--no-ocr" in sys.argv
    skip_individual = "--no-pdf" in sys.argv

    print("=" * 60)
    print("  CS-371 ANTALYA — 3 KAYNAK KARŞILAŞTIRMASI")
    print("=" * 60)
    if skip_ocr:
        print("  [--no-ocr] Toplu PDF atlanıyor")

    init_db()
    reload_mapping()
    clear_missing_clubs()

    results_map = {}

    # KAYNAK 0: Lenex
    r0 = kaynak_lenex()
    results_map["0: Lenex"] = r0
    if r0:
        export(r0, "0_Lenex", "Lenex")

    # KAYNAK A: Bireysel PDF'ler
    if not skip_individual:
        rA = kaynak_bireysel_pdf()
        results_map["A: Bireysel PDF"] = rA
        if rA:
            export(rA, "A_BireyselPDF", "Bireysel PDF")
    else:
        results_map["A: Bireysel PDF"] = []

    # KAYNAK B: Toplu PDF (OCR)
    if not skip_ocr:
        rB = kaynak_toplu_pdf()
        results_map["B: Toplu PDF (OCR)"] = rB
        if rB:
            export(rB, "B_TopluPDF_OCR", "Toplu PDF (OCR)")
    else:
        results_map["B: Toplu PDF (OCR)"] = []

    # Karşılaştırma
    compare(results_map)

    # Federasyon karma sıralaması (Antalya)
    print("\n" + "═"*60)
    print("FED. KARMA SIRALAMASI")
    print("═"*60)
    export_karma_rankings()

    print("\n✓ Tamamlandı.")
    print(f"  Yarış Excel:  {OUT_DIR}")
    print(f"  Karma Excel:  {KARMA_DIR}")
