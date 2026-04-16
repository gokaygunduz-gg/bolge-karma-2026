"""
database/sync_mapping.py — Excel Mapping → SQLite Sync

Excel'deki Kulüp–Şehir–Bölge verisini SQLite'a kopyalar.
Bu işlem SADECE Excel'den DB'ye doğru çalışır — DB'den Excel'e hiçbir şey yazılmaz.

Ne zaman çalıştırılır:
  - İlk kurulumda (DB boş)
  - Kullanıcı Excel'e yeni kulüp eklediğinde

Kullanım:
  from database.sync_mapping import sync_clubs_from_excel

  stats = sync_clubs_from_excel()
  print(stats)  # {'loaded': 2033, 'skipped': 197, 'duration_s': 1.4}

Komut satırından:
  cd "Bölge Karmaları 2026"
  python -m database.sync_mapping
"""

import time
import logging
import openpyxl

from config import (
    MAPPING_EXCEL_PATH,
    MAPPING_SHEET_NAME,
    COL_CLUB_ALT,
    COL_CLUB_CANONICAL,
    COL_CITY,
    COL_REGION,
    MAPPING_DATA_START_ROW,
)
from modules.m1_normalize import normalize_for_lookup, restore_turkish_club
from database.db import get_connection, upsert_club, init_db

logger = logging.getLogger(__name__)


def sync_clubs_from_excel(verbose: bool = True) -> dict:
    """
    Excel'deki tüm kulüp satırlarını DB'ye yazar (UPSERT).

    Döndürür:
      {
        'loaded':     int,  — başarıyla işlenen satır sayısı
        'skipped':    int,  — atlanan satır (eksik veri)
        'duration_s': float — işlem süresi
      }
    """
    start = time.time()
    loaded = 0
    skipped = 0

    if verbose:
        print(f"Excel okunuyor: {MAPPING_EXCEL_PATH}")

    try:
        wb = openpyxl.load_workbook(MAPPING_EXCEL_PATH, read_only=True, data_only=True)
    except FileNotFoundError:
        logger.error("Mapping Excel bulunamadı: %s", MAPPING_EXCEL_PATH)
        return {"loaded": 0, "skipped": 0, "duration_s": 0, "error": "Dosya bulunamadı"}

    if MAPPING_SHEET_NAME not in wb.sheetnames:
        wb.close()
        return {"loaded": 0, "skipped": 0, "duration_s": 0, "error": f"Sheet yok: {MAPPING_SHEET_NAME}"}

    ws = wb[MAPPING_SHEET_NAME]
    conn = get_connection()

    # Toplu insert için transaction
    with conn:
        for row in ws.iter_rows(min_row=MAPPING_DATA_START_ROW, values_only=True):
            col_a = row[COL_CLUB_ALT]        if len(row) > COL_CLUB_ALT        else None
            col_c = row[COL_CLUB_CANONICAL]  if len(row) > COL_CLUB_CANONICAL  else None
            col_e = row[COL_CITY]            if len(row) > COL_CITY            else None
            col_g = row[COL_REGION]          if len(row) > COL_REGION          else None

            # Şehir veya Bölge yoksa atla
            if not col_e or col_g is None:
                skipped += 1
                continue

            # En az bir kulüp ismi olmalı
            if not col_a and not col_c:
                skipped += 1
                continue

            try:
                region = int(col_g)
            except (ValueError, TypeError):
                skipped += 1
                continue

            city      = str(col_e).strip()
            canonical = restore_turkish_club(str(col_c).strip()) if col_c else None
            alt       = str(col_a).strip() if col_a else None

            # Arama anahtarı: önce alternatif isim, yoksa kanonik
            primary_name = alt or canonical
            normalized   = normalize_for_lookup(primary_name)

            if not normalized:
                skipped += 1
                continue

            upsert_club(
                name_alt        = primary_name,
                name_canonical  = canonical,
                name_normalized = normalized,
                city            = city,
                region          = region,
            )

            # Kanonik isim de ayrıca ekle (farklı normalize değere sahipse)
            if canonical:
                canonical_normalized = normalize_for_lookup(canonical)
                if canonical_normalized and canonical_normalized != normalized:
                    upsert_club(
                        name_alt        = canonical,
                        name_canonical  = canonical,
                        name_normalized = canonical_normalized,
                        city            = city,
                        region          = region,
                    )

            loaded += 1

    wb.close()

    # Sync logunu kaydet
    duration = round(time.time() - start, 2)
    conn.execute("""
        INSERT INTO sync_log (source_file, rows_loaded, rows_skipped, notes)
        VALUES (?, ?, ?, ?)
    """, (MAPPING_EXCEL_PATH, loaded, skipped, f"Süre: {duration}s"))
    conn.commit()

    stats = {"loaded": loaded, "skipped": skipped, "duration_s": duration}

    if verbose:
        print(f"  Sync tamamlandı: {loaded} kulüp yüklendi, {skipped} satır atlandı ({duration}s)")

    return stats


def is_mapping_synced() -> bool:
    """
    DB'de kulüp verisi var mı? (En az 1 kayıt)
    Uygulama başlangıcında kontrol için kullanılır.
    """
    conn = get_connection()
    count = conn.execute("SELECT COUNT(*) FROM clubs").fetchone()[0]
    return count > 0


def ensure_mapping_synced(verbose: bool = True) -> None:
    """
    DB'de mapping yoksa otomatik sync yapar.
    Uygulama başlangıcında çağrılır.
    """
    if not is_mapping_synced():
        if verbose:
            print("DB'de mapping bulunamadı. Excel'den otomatik yükleniyor...")
        sync_clubs_from_excel(verbose=verbose)
    elif verbose:
        conn = get_connection()
        count = conn.execute("SELECT COUNT(*) FROM clubs").fetchone()[0]
        print(f"  Mapping zaten DB'de: {count} kulüp")


# ─────────────────────────────────────────────────────────────────────────────
# Komut satırından çalıştırma
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    import os
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

    logging.basicConfig(level=logging.WARNING)
    init_db()
    stats = sync_clubs_from_excel(verbose=True)
    print(f"\nSonuç: {stats}")
