"""
export/excel_exporter.py — Yarış Sonuçlarını Excel'e Aktar

Format (eski projeyle aynı — 24 sütun):
  Sütun 1-7  : ad_soyad | yb | kulüp | şehir | bölge | yaş | cinsiyet
  Sütun 8-24 : 17 yarış sütunu (Serbest_50m … Karışık_400m)

Her satır = 1 sporcu (en iyi süre önceden _dedup_best_time ile alınmış olmalı).
Yarışılmayan branşlar NaN/boş kalır.

Kullanım:
  from export.excel_exporter import export_race_results

  path = export_race_results(results, "cs_370_sonuclari.xlsx")
  print(f"Kaydedildi: {path}")
"""

import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import BASE_COLUMNS, FIXED_EVENT_COLUMNS, ALL_COLUMNS, OUTPUT_RACE_DIR

# ─────────────────────────────────────────────────────────────────────────────
# Dahili sabitler
# ─────────────────────────────────────────────────────────────────────────────

# result dict anahtarları → Excel sütun başlığı (BASE_COLUMNS sırası)
_BASE_HEADER = ["Ad Soyad", "YB", "Kulüp", "Şehir", "Bölge", "Yaş", "Cinsiyet"]

# Stil adı + mesafe → FIXED_EVENT_COLUMNS'daki sütun adı
# Örnek: ("Serbest", 50) → "Serbest_50m"
_EVENT_COL: dict[tuple, str] = {}
for _col in FIXED_EVENT_COLUMNS:
    # "Serbest_50m" → stroke="Serbest", dist=50
    _parts = _col.rsplit("_", 1)
    _stroke = _parts[0]
    _dist   = int(_parts[1].rstrip("m"))
    _EVENT_COL[(_stroke, _dist)] = _col

# Stil sütun başlıkları: "50m" → birden fazla stilde aynı mesafe
_EVENT_HEADER = FIXED_EVENT_COLUMNS  # sütun başlığı = kolon adı

# ─────────────────────────────────────────────────────────────────────────────
# Stil / renk sabitleri
# ─────────────────────────────────────────────────────────────────────────────

_HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")   # Koyu lacivert
_HEADER_FONT   = Font(color="FFFFFF", bold=True, size=10)
_SUBHDR_FILL   = PatternFill("solid", fgColor="2E75B6")   # Açık mavi (yarış sütunları)
_SUBHDR_FONT   = Font(color="FFFFFF", bold=True, size=9)
_ALT_FILL      = PatternFill("solid", fgColor="F2F2F2")   # Gri alternatif satır
_BORDER_THIN   = Border(
    left   = Side(style="thin", color="D0D0D0"),
    right  = Side(style="thin", color="D0D0D0"),
    top    = Side(style="thin", color="D0D0D0"),
    bottom = Side(style="thin", color="D0D0D0"),
)

_GENDER_MAP = {"M": "Erkek", "F": "Kız"}


# ─────────────────────────────────────────────────────────────────────────────
# Pivot oluştur
# ─────────────────────────────────────────────────────────────────────────────

def _build_pivot(results: list[dict]) -> list[dict]:
    """
    Düz sonuç listesini (1 satır = 1 yarış) pivot'a (1 satır = 1 sporcu) çevirir.

    Gruplandırma anahtarı: (name_normalized, matched_club_normalized)
    Eşleşmiş kulüp adı kullanılır, böylece aynı kulübün farklı OCR varyantları
    ("CK S.K." / "Ck Spor Kulibu" gibi) aynı satırda birleşir.
    Her yarış branşı için en iyi süre alınır (zaten _dedup_best_time yapılmış olmalı,
    ama burada da güvenlik katmanı olarak min süre alınır).
    """
    from modules.m1_normalize import normalize_for_lookup

    # Sporcu → bilgi + event_times
    athletes: dict[tuple, dict] = {}

    for r in results:
        # Matched canonical kulüp adı — OCR varyantlarını birleştirir
        club_for_key = r.get("club") or r.get("club_raw") or ""
        # Display name kullan (override sonrası) — aynı sporcu farklı OCR formlarıyla
        # geldiğinde aynı satırda birleşsin
        key = (
            normalize_for_lookup(r["name"]),
            normalize_for_lookup(club_for_key),
        )
        if key not in athletes:
            athletes[key] = {
                "ad_soyad":  r["name"],
                "yb":        r["yb"],
                "kulüp":     r["club"],
                "şehir":     r["city"] or "",
                "bölge":     r["region"] or "",
                "yaş":       r["age"] or "",
                "cinsiyet":  _GENDER_MAP.get(r["gender"], r["gender"]),
                "_times":    {},   # col_name → (time_text, time_seconds)
            }
        else:
            # Daha iyi bilgi varsa güncelle (matched kulüp, şehir, bölge)
            existing_rec = athletes[key]
            if not existing_rec["şehir"] and r.get("city"):
                existing_rec["şehir"] = r["city"]
                existing_rec["bölge"] = r.get("region") or ""
                existing_rec["kulüp"] = r["club"]

        event_col = _EVENT_COL.get((r["stroke"], r["distance"]))
        if event_col is None:
            continue  # Bilinmeyen branş — atla

        existing = athletes[key]["_times"].get(event_col)
        if existing is None or r["time_seconds"] < existing[1]:
            athletes[key]["_times"][event_col] = (r["time_text"], r["time_seconds"])

    # Sıralama: şehir → kulüp → ad
    rows = list(athletes.values())
    rows.sort(key=lambda x: (x["şehir"] or "ZZZ", x["kulüp"] or "ZZZ", x["ad_soyad"]))
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Excel yazma
# ─────────────────────────────────────────────────────────────────────────────

def _write_sheet(ws, pivot_rows: list[dict], title: str = "") -> None:
    """Pivot satırlarını verilen worksheet'e yazar."""

    # ── Başlık satırı (1. satır)
    headers = _BASE_HEADER + _EVENT_HEADER
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        if col_idx <= len(_BASE_HEADER):
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
        else:
            cell.fill = _SUBHDR_FILL
            cell.font = _SUBHDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER_THIN

    # ── Veri satırları (2. satırdan itibaren)
    for row_idx, athlete in enumerate(pivot_rows, start=2):
        fill = _ALT_FILL if row_idx % 2 == 0 else None

        base_values = [
            athlete["ad_soyad"],
            athlete["yb"],
            athlete["kulüp"],
            athlete["şehir"],
            athlete["bölge"],
            athlete["yaş"],
            athlete["cinsiyet"],
        ]

        for col_idx, val in enumerate(base_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if fill:
                cell.fill = fill
            cell.border = _BORDER_THIN
            cell.alignment = Alignment(vertical="center")

        for col_idx, event_col in enumerate(FIXED_EVENT_COLUMNS, start=len(_BASE_HEADER) + 1):
            time_entry = athlete["_times"].get(event_col)
            val = time_entry[0] if time_entry else ""
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if fill:
                cell.fill = fill
            cell.border = _BORDER_THIN
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Sütun genişlikleri
    col_widths = [28, 5, 32, 14, 7, 5, 9] + [9] * len(FIXED_EVENT_COLUMNS)
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Satır yüksekliği (başlık)
    ws.row_dimensions[1].height = 30

    # ── Pencereyi dondur (başlık + 7 temel sütun)
    ws.freeze_panes = "H2"

    # ── Sayfa başlığı
    if title:
        ws.title = title[:31]  # Excel sheet adı max 31 karakter


# ─────────────────────────────────────────────────────────────────────────────
# Ana export fonksiyonu
# ─────────────────────────────────────────────────────────────────────────────

def export_race_results(
    results:   list[dict],
    filename:  str | None = None,
    title:     str        = "Sonuçlar",
) -> str:
    """
    Yarış sonuçlarını Excel dosyasına aktar.

    Parametreler:
      results:  scrape_race() çıktısı (zenginleştirilmiş dict listesi)
      filename: Dosya adı (None → otomatik timestamp oluşturulur)
      title:    Sheet ve dosya başlığı

    Döndürür: Oluşturulan dosyanın tam yolu
    """
    os.makedirs(OUTPUT_RACE_DIR, exist_ok=True)

    race_date = results[0].get("race_date", "") if results else ""

    if filename is None:
        ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Yaris_Sonuclari_{ts}.xlsx"

    if not filename.endswith(".xlsx"):
        filename += ".xlsx"

    # race_date öneki ekle (henüz eklenmemişse)
    if race_date and not filename.startswith(race_date):
        filename = f"{race_date}_{filename}"

    output_path = os.path.join(OUTPUT_RACE_DIR, filename)

    pivot_rows = _build_pivot(results)

    wb = openpyxl.Workbook()
    ws = wb.active
    _write_sheet(ws, pivot_rows, title=title)

    wb.save(output_path)
    return output_path
