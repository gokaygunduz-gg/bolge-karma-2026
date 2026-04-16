"""Antalya-Edirne sporcu karşılaştırma Excel'i oluşturur."""
import sqlite3, openpyxl, re, sys, unicodedata, urllib.request
from openpyxl.styles import Font, PatternFill, Alignment
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import fitz  # pymupdf

def tr_norm(s):
    """Turkish-aware normalization: İ/I/ı/i→i, strip accents, normalize whitespace."""
    if not s: return ""
    s = s.replace("İ", "I").replace("ı", "i")
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower()
    s = " ".join(s.split())   # collapse multiple spaces
    return s.strip()

# Suffixes to strip for club comparison
_CLUB_STRIP = re.compile(
    r"\b(spor\s+kulubu?|belediyesi?|buyuksehir|genclik\s+ve\s+spor\s+i[li]\s+mudurlugu?|"
    r"il\s+mudurlugu?|spor\s+kulubu?|a\.s\.?|as\b)\b",
    re.IGNORECASE
)

def club_key(club):
    """Extra normalization for club comparison: strip generic suffixes + normalize."""
    s = tr_norm(club)
    s = _CLUB_STRIP.sub("", s)
    s = " ".join(s.split())
    return s

def name_key(name, by):
    return (tr_norm(name), int(by) if by else 0)

# ── Parse Edirne ProgressionDetails ─────────────────────────────────────────
url = "https://canli.tyf.gov.tr/tyf/cs-376/ProgressionDetails.pdf"
req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
with urllib.request.urlopen(req, timeout=20) as r:
    data = r.read()
doc = fitz.open(stream=data, filetype="pdf")
lines = []
for page in doc:
    lines += [l.rstrip() for l in page.get_text().split("\n")]

STROKE_NAMES = ["Serbest","Sırtüstü","Kurbağalama","Kelebek","Karışık","Prelim","Final","Barajlar","Timed"]
HEADER_PARTS = ["Splash Meet Manager","Registered to Turkiye","TÜRKİYE YILDIZ","EDİRNE, 17",
                "Sporcuların İlerlemesi","Tüm Yarışlar","Sayfa"]
DIST_RE = re.compile(r"^\d{2,4}m\s")
TIME_RE = re.compile(r"^(\d+:)?\d+\.\d{2}$|^NT$|^DSQ$|^DNS$|^DQ$|^-$")
DATE_RE = re.compile(r"^\d{2}\.\d{2}\.\d{4}")
ANY_ATHLETE_RE = re.compile(r"^.+?,\s+\d{4}\s+\(\d+\s+ya[sş]\),\s+(Erkekler|Bayanlar|Kızlar|Erkek|Kız)")
TARGET_ATHLETE_RE = re.compile(r"^(.+?),\s+(\d{4})\s+\(\d+\s+ya[sş]\),\s+(Erkekler|Bayanlar|Kızlar|Erkek|Kız)$")
TARGET_BYS = {2011, 2012, 2013}
GENDER_MAP = {"Erkekler":"M","Erkek":"M","Bayanlar":"F","Kızlar":"F","Kız":"F"}

def is_skip(line):
    if not line or line == "-": return True
    if TIME_RE.match(line) or DATE_RE.match(line) or DIST_RE.match(line): return True
    if any(kw in line for kw in STROKE_NAMES + HEADER_PARTS): return True
    if ANY_ATHLETE_RE.match(line): return True
    return False

edirne_athletes = {}
current_club = ""
for lr in lines:
    line = lr.strip()
    m = TARGET_ATHLETE_RE.match(line)
    if m:
        name_raw, by_s, gr = m.group(1).strip(), m.group(2), m.group(3)
        by = int(by_s)
        gender = GENDER_MAP.get(gr, "")
        if by in TARGET_BYS and gender:
            key = name_key(name_raw, by)
            if key not in edirne_athletes:
                edirne_athletes[key] = {"name": name_raw, "birth_year": by, "gender": gender, "club": current_club}
        continue
    if line and not is_skip(line):
        current_club = line

# ── Antalya DB ────────────────────────────────────────────────────────────────
conn = sqlite3.connect("data/bolge_karmalari.db")
conn.row_factory = sqlite3.Row
rows = conn.execute(
    "SELECT DISTINCT athlete_name, birth_year, gender, club FROM fed_results "
    "WHERE race_leg=? AND birth_year IN (2011,2012,2013) ORDER BY athlete_name",
    ("antalya",)
).fetchall()
conn.close()

antalya_athletes = {}
for r in rows:
    key = name_key(r["athlete_name"], r["birth_year"])
    if key not in antalya_athletes:
        antalya_athletes[key] = {
            "name": r["athlete_name"], "birth_year": r["birth_year"],
            "gender": r["gender"], "club": r["club"]
        }

print(f"Antalya unique (normalized): {len(antalya_athletes)}")
print(f"Edirne unique (normalized):  {len(edirne_athletes)}")

# ── Merge ─────────────────────────────────────────────────────────────────────
all_keys = set(antalya_athletes.keys()) | set(edirne_athletes.keys())
sorted_keys = sorted(all_keys, key=lambda k: (k[0], k[1]))

def club_match(a, b):
    if not a or not b: return True
    # Level 1: exact after full normalization
    na, nb = tr_norm(a), tr_norm(b)
    if na == nb or na in nb or nb in na:
        return True
    # Level 2: strip generic suffixes, then compare
    ka, kb = club_key(a), club_key(b)
    if ka and kb and (ka == kb or ka in kb or kb in ka):
        return True
    return False

FILL_BOTH      = PatternFill("solid", fgColor="E8F5E9")
FILL_ANT_ONLY  = PatternFill("solid", fgColor="FFF3CD")
FILL_ED_ONLY   = PatternFill("solid", fgColor="D1ECF1")
FILL_CLUB_DIFF = PatternFill("solid", fgColor="FADBD8")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sporcu Karşılaştırma"

HEADER = ["Antalya - Sporcu","Antalya - Kulüp","Doğum Yılı","Cinsiyet",
          "Edirne - Sporcu","Edirne - Kulüp","",""]
hfont = Font(bold=True, color="FFFFFF", size=11)
hfill = PatternFill("solid", fgColor="1F3864")
for ci, h in enumerate(HEADER, 1):
    cell = ws.cell(1, ci, h)
    cell.font = hfont
    cell.fill = hfill
    cell.alignment = Alignment(horizontal="center", vertical="center")

stats = {"both": 0, "ant_only": 0, "ed_only": 0, "club_diff": 0}

for ri, key in enumerate(sorted_keys, 2):
    ant = antalya_athletes.get(key)
    ed  = edirne_athletes.get(key)
    has_ant = bool(ant)
    has_ed  = bool(ed)
    if has_ant and has_ed:
        if club_match(ant["club"], ed["club"]):
            fill = FILL_BOTH
            stats["both"] += 1
        else:
            fill = FILL_CLUB_DIFF
            stats["club_diff"] += 1
    elif has_ant:
        fill = FILL_ANT_ONLY
        stats["ant_only"] += 1
    else:
        fill = FILL_ED_ONLY
        stats["ed_only"] += 1

    base = ant if ant else ed
    ant_name = ant["name"] if ant else ""
    ant_club = ant["club"] if ant else ""
    ed_name  = ed["name"]  if ed  else ""
    ed_club  = ed["club"]  if ed  else ""
    by = base["birth_year"]
    g  = base["gender"]

    for ci, val in enumerate([ant_name, ant_club, by, g, ed_name, ed_club, "", ""], 1):
        cell = ws.cell(ri, ci, val)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")

ws.column_dimensions["A"].width = 32
ws.column_dimensions["B"].width = 48
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 10
ws.column_dimensions["E"].width = 32
ws.column_dimensions["F"].width = 48
ws.row_dimensions[1].height = 28
ws.freeze_panes = "A2"

ls = wb.create_sheet("Açıklama")
ls["A1"] = "Renk"; ls["A1"].font = Font(bold=True)
ls["B1"] = "Anlam"; ls["B1"].font = Font(bold=True)
legend = [
    ("Yeşil", FILL_BOTH,      "Her iki yarışta da var, kulüp eşleşiyor"),
    ("Kırmızı", FILL_CLUB_DIFF, "Her iki yarışta da var, KULÜP FARKLI"),
    ("Sarı",  FILL_ANT_ONLY,  "Sadece Antalya (Edirne'ye gelmiyor)"),
    ("Mavi",  FILL_ED_ONLY,   "Sadece Edirne (yeni sporcu)"),
]
for i, (c, f, d) in enumerate(legend, 2):
    ls.cell(i, 1, c).fill = f
    ls.cell(i, 2, d)
ls.column_dimensions["A"].width = 12
ls.column_dimensions["B"].width = 55

out = "data/antalya_edirne_sporcu_karsilastirma.xlsx"
wb.save(out)
print(f"Excel kaydedildi: {out}")
print(f"  Toplam satır: {len(sorted_keys)}")
print(f"  Her ikisinde de (yeşil): {stats['both']}")
print(f"  Sadece Antalya  (sarı):  {stats['ant_only']}")
print(f"  Sadece Edirne   (mavi):  {stats['ed_only']}")
print(f"  Kulüp farklı  (kırmızı): {stats['club_diff']}")
